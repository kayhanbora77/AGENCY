import polars as pl
import duckdb
import re
import sys
from datetime import datetime
from pathlib import Path

INPUT_FILE = "/home/kayhan/Desktop/Gelen_Datalar/NEW_DATA_MURAT/GCC/GCC 2025 (C2R).xlsx"
DATABASE_DIR = Path.home() / "my_database"
DATABASE_NAME = "my_db.duckdb"
DB_PATH = DATABASE_DIR / DATABASE_NAME

TABLE_NAME = "GCC"
CHUNK_SIZE = 100_000

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

MONTH_MAP = {
    m: i
    for i, m in enumerate(
        [
            "JAN",
            "FEB",
            "MAR",
            "APR",
            "MAY",
            "JUN",
            "JUL",
            "AUG",
            "SEP",
            "OCT",
            "NOV",
            "DEC",
        ],
        1,
    )
}


def resolve_short_date(token: str, year: int) -> str | None:
    """Convert '06JUN' -> '2025-06-06' using supplied year."""
    m = re.match(r"^(\d{1,2})([A-Z]{3})$", token.strip().upper())
    if m:
        day, mon = int(m.group(1)), MONTH_MAP.get(m.group(2))
        if mon:
            return f"{year}-{mon:02d}-{day:02d}"
    return None


def parse_any_date(val, fallback_year: int | None = None) -> str | None:
    if val is None:
        return None
    # Handle datetime/date objects from openpyxl directly
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    if not str(val).strip() or str(val).strip() in ("", "NULL", "None"):
        return None
    val = str(val).strip()
    # Already ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}$", val):
        return val
    # "25-Jan-25" or "25-Jan-2025"
    m = re.match(r"^(\d{1,2})-([A-Za-z]{3})-(\d{2,4})$", val)
    if m:
        day, mon_str, yr = int(m.group(1)), m.group(2).upper(), m.group(3)
        mon = MONTH_MAP.get(mon_str)
        if mon:
            year = int(yr) + (2000 if len(yr) == 2 else 0)
            return f"{year}-{mon:02d}-{day:02d}"
    # "6-Jun-25" pandas-style
    # Short token (e.g. "06JUN") — must check BEFORE dateutil which guesses wrong year
    if fallback_year:
        r = resolve_short_date(val, fallback_year)
        if r:
            return r
    try:
        from dateutil import parser as dparser

        return dparser.parse(val, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        pass
    return None


def fix_ftda(ftda_val, first_sector_date) -> list[str]:
    """
    FTDA can be '06JUN 07JUN ...' or space-separated dates.
    Returns a list of ISO date strings, one per token.
    Year is resolved from FirstSectordate.
    """
    # If openpyxl gave us a real datetime, wrap it directly
    if hasattr(ftda_val, "strftime"):
        return [ftda_val.strftime("%Y-%m-%d")]
    if not ftda_val or str(ftda_val).strip() in ("", "NULL", "None"):
        return []
    fallback_year = None
    if first_sector_date:
        d = parse_any_date(first_sector_date)
        if d:
            fallback_year = int(d[:4])
    results = []
    for t in str(ftda_val).split():
        t = t.strip()
        if not t:
            continue
        result = parse_any_date(t, fallback_year)
        if result:
            results.append(result)
    return results


def parse_flights(val: str) -> list[str]:
    """'DL-9698 DL-9320   ' -> ['DL9698', 'DL9320']"""
    if not val or val.strip() in ("", "NULL"):
        return []
    parts = val.split()
    out = []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        clean = re.sub(r"([A-Z0-9]{2})-(\d+)", r"\1\2", p)
        out.append(clean)
    return out


def parse_airports(val: str) -> list[str]:
    """'YYZ/AMS AMS/DXB' -> ['YYZ', 'AMS', 'DXB'] (unique, ordered)"""
    if not val or val.strip() in ("", "NULL"):
        return []
    seen, out = set(), []
    for part in val.split():
        for code in part.split("/"):
            code = code.strip()
            if code and code not in seen:
                seen.add(code)
                out.append(code)
    return out


# ---------------------------------------------------------------------------
# Determine max flights / airports by scanning first sheet sample
# ---------------------------------------------------------------------------


def scan_max_widths(df: pl.DataFrame):
    max_f = max_a = max_ftda = 0
    cols = ["FlightNo", "Sector", "FTDA"]
    for row in df.select(cols).iter_rows():
        f, s, ftda = row
        max_f = max(max_f, len(parse_flights(str(f) if f else "")))
        max_a = max(max_a, len(parse_airports(str(s) if s else "")))
        max_ftda = max(max_ftda, len(fix_ftda(str(ftda) if ftda else "", None)))
    return max_f, max_a, max_ftda


# ---------------------------------------------------------------------------
# Transform one batch (list of dicts) -> list of dicts
# ---------------------------------------------------------------------------


def transform_rows(
    rows: list[dict], max_flights: int, max_airports: int, max_ftda: int
) -> list[dict]:
    out = []
    for r in rows:
        # FTDA resolved first, before FirstSectordate is overwritten
        ftda_dates = fix_ftda(r.get("FTDA"), r.get("FirstSectordate"))
        for i in range(max_ftda):
            r[f"FTDA{i + 1}"] = ftda_dates[i] if i < len(ftda_dates) else None
        r.pop("FTDA", None)

        # Date columns (parse_any_date handles raw datetime objects and string formats)
        for col in ("DAIS", "FirstSectordate", "LastSectordate"):
            r[col] = parse_any_date(r.get(col))

        # FlightNo -> FlightNo1 .. FlightNoN
        flights = parse_flights(str(r.get("FlightNo", "") or ""))
        for i in range(max_flights):
            r[f"FlightNo{i + 1}"] = flights[i] if i < len(flights) else None
        r.pop("FlightNo", None)

        # Sector -> Airport1 .. AirportN
        airports = parse_airports(str(r.get("Sector", "") or ""))
        for i in range(max_airports):
            r[f"Airport{i + 1}"] = airports[i] if i < len(airports) else None
        r.pop("Sector", None)

        out.append(r)
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

import openpyxl

print(f"Opening workbook: {INPUT_FILE}")
wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
sheet_names = wb.sheetnames
print(f"Sheets found: {sheet_names}")
wb.close()

con = duckdb.connect(DB_PATH)
con.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")
table_created = False

for sheet in sheet_names:
    print(f"\nProcessing sheet: {sheet}")

    # Read entire sheet with polars (lazy via openpyxl row streaming)
    # We use pandas engine as bridge but in chunks via openpyxl directly
    wb2 = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb2[sheet]

    rows_iter = ws.iter_rows(values_only=True)
    headers = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(next(rows_iter))
    ]

    # First pass: scan 5000 rows to find max flights/airports
    sample_rows = []
    for _, row in zip(range(5000), rows_iter):
        sample_rows.append(dict(zip(headers, row)))

    sample_df = pl.DataFrame(
        {
            h: [str(r.get(h)) if r.get(h) is not None else None for r in sample_rows]
            for h in headers
        },
        infer_schema_length=500,
        strict=False,
    )

    max_f, max_a, max_ftda = scan_max_widths(sample_df)
    print(f"  Max flights: {max_f}, Max airports: {max_a}, Max FTDA dates: {max_ftda}")

    # Process sample + continue streaming
    batch = sample_rows
    total = 0

    def flush_batch(batch, first=False):
        global table_created
        transformed = transform_rows(batch, max_f, max_a, max_ftda)
        # Stringify all values to avoid mixed-type errors from openpyxl datetime objects
        str_transformed = [
            {k: (str(v) if v is not None else None) for k, v in row.items()}
            for row in transformed
        ]
        pdf = pl.DataFrame(
            str_transformed, infer_schema_length=len(str_transformed), strict=False
        )
        # Cast all date cols to string (they're already ISO strings or None)
        if not table_created:
            con.execute(f"CREATE TABLE {TABLE_NAME} AS SELECT * FROM pdf LIMIT 0")
            table_created = True
        con.execute(f"INSERT INTO {TABLE_NAME} SELECT * FROM pdf")

    while batch:
        flush_batch(batch)
        total += len(batch)
        print(f"  Inserted {total:,} rows...", end="\r")
        batch = []
        chunk = []
        for row in rows_iter:
            chunk.append(dict(zip(headers, row)))
            if len(chunk) >= CHUNK_SIZE:
                break
        batch = chunk

    print(f"  Sheet '{sheet}' done — {total:,} rows total.")
    wb2.close()

count = con.execute(f"SELECT COUNT(*) FROM {TABLE_NAME}").fetchone()[0]
cols = con.execute(f"PRAGMA table_info({TABLE_NAME})").fetchdf()
print(f"\n✅ Done! Table '{TABLE_NAME}' in '{DB_PATH}'")
print(f"   Total rows : {count:,}")
print(f"   Columns    : {len(cols)}")
print(f"\nColumn list:\n{cols[['name', 'type']].to_string(index=False)}")

con.close()
