import openpyxl
import duckdb
import re
import uuid
import polars as pl
from pathlib import Path

INPUT_FILE = "/home/kayhan/Desktop/Gelen_Datalar/NEW_DATA_MURAT/INTL/Intl_ data.xlsx"
DATABASE_DIR = Path.home() / "my_database"
DATABASE_NAME = "my_db.duckdb"
DB_PATH = DATABASE_DIR / DATABASE_NAME

TABLE_NAME = "INTL"
CHUNK_SIZE = 100_000

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def fix_flight_number(val) -> str | None:
    """
    'BA 119 / BA 198'  -> 'BA119'
    'EK 506'           -> 'EK506'
    """
    if val is None:
        return None
    v = str(val).strip()
    if not v or v in ("NULL", "None", ""):
        return None
    first = v.split("/")[0].strip()
    return first.replace(" ", "") or None


def parse_sector(val):
    """
    'LHR - BLR / BOM - LHR'  -> Airport1='LHR', Airport2='BLR'
    'DXB - BOM'               -> Airport1='DXB', Airport2='BOM'
    'LHR - BOM'               -> Airport1='LHR', Airport2='BOM'
    """
    if val is None:
        return None, None
    v = str(val).strip()
    if not v or v in ("NULL", "None", ""):
        return None, None
    first_segment = v.split("/")[0].strip()  # e.g. "LHR - BLR"
    parts = [p.strip() for p in first_segment.split("-")]
    airport1 = parts[0].strip() if len(parts) > 0 else None
    airport2 = parts[1].strip() if len(parts) > 1 else None
    return airport1 or None, airport2 or None


def safe_str(val) -> str | None:
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    v = str(val).strip()
    return None if v in ("", "NULL", "None") else v


# ---------------------------------------------------------------------------
# Transform one batch
# ---------------------------------------------------------------------------


def transform_batch(
    batch: list[dict],
    headers: list[str],
    flightno_col: str | None,
    sector_col: str | None,
) -> pl.DataFrame:
    # Output columns: all original cols EXCEPT sector_col,
    # with FLIGHTNO cleaned, then Airport1 + Airport2 + Id appended.
    rows = []
    for r in batch:
        row = {}
        for k in headers:
            if k == sector_col:
                continue  # drop SECTOR — replaced below
            v = r.get(k)
            if k == flightno_col:
                row[k] = fix_flight_number(v)  # clean in-place, keep col name
            else:
                row[k] = safe_str(v)

        a1, a2 = parse_sector(r.get(sector_col))
        row["Airport1"] = a1
        row["Airport2"] = a2
        row["Id"] = str(uuid.uuid4())
        rows.append(row)

    return pl.DataFrame(rows, infer_schema_length=len(rows), strict=False)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

DATABASE_DIR.mkdir(parents=True, exist_ok=True)

print(f"Opening workbook: {INPUT_FILE}")
wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
sheet_names = wb.sheetnames
print(f"Sheets found: {sheet_names}")
wb.close()

con = duckdb.connect(DB_PATH)
con.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")
table_created = False
grand_total = 0

for sheet in sheet_names:
    print(f"\nProcessing sheet: {sheet}")
    wb2 = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb2[sheet]
    rows_iter = ws.iter_rows(values_only=True)

    raw_headers = next(rows_iter)
    headers = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(raw_headers)
    ]

    # Case-insensitive exact match for FLIGHTNO and SECTOR
    hl = {h.strip().upper(): h for h in headers}
    flightno_col = hl.get("FLIGHTNO")
    sector_col = hl.get("SECTOR")

    print(f"  All columns      : {headers}")
    print(f"  FLIGHTNO col     : {flightno_col}")
    print(f"  SECTOR col       : {sector_col}")

    sheet_total = 0
    batch = []

    def flush(batch):
        global table_created
        if not batch:
            return
        pdf = transform_batch(batch, headers, flightno_col, sector_col)
        if not table_created:
            con.execute(f"CREATE TABLE {TABLE_NAME} AS SELECT * FROM pdf LIMIT 0")
            table_created = True
        con.execute(f"INSERT INTO {TABLE_NAME} SELECT * FROM pdf")

    for row in rows_iter:
        if all(v is None for v in row):
            continue
        batch.append(dict(zip(headers, row)))
        if len(batch) >= CHUNK_SIZE:
            flush(batch)
            sheet_total += len(batch)
            print(f"  Inserted {sheet_total:,} rows...", end="\r")
            batch = []

    if batch:
        flush(batch)
        sheet_total += len(batch)

    print(f"  Sheet '{sheet}' done — {sheet_total:,} rows.")
    grand_total += sheet_total
    wb2.close()

count = con.execute(f"SELECT COUNT(*) FROM {TABLE_NAME}").fetchone()[0]
cols = con.execute(f"PRAGMA table_info({TABLE_NAME})").fetchdf()

print(f"\n✅ Done! Table '{TABLE_NAME}' in '{DB_PATH}'")
print(f"   Total rows : {count:,}")
print(f"   Columns    : {len(cols)}")
print(f"\n{cols[['name', 'type']].to_string(index=False)}")

# Sanity check
print("\nSample FLIGHTNO / Airport1 / Airport2:")
print(
    con.execute(f"SELECT FLIGHTNO, Airport1, Airport2, Id FROM {TABLE_NAME} LIMIT 5")
    .fetchdf()
    .to_string(index=False)
)

con.close()
