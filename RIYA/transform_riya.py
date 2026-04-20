import openpyxl
import polars as pl
import duckdb
import re
import sys
from datetime import datetime
from pathlib import Path

INPUT_FILE = (
    "/home/kayhan/Desktop/Gelen_Datalar/NEW_DATA_MURAT/RIYA/Riya 25-26 Raw Data.xlsx"
)
DATABASE_DIR = Path.home() / "my_database"
DATABASE_NAME = "my_db.duckdb"
DB_PATH = DATABASE_DIR / DATABASE_NAME

TABLE_NAME = "RIYA"
CHUNK_SIZE = 100_000


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def parse_date(val) -> str | None:
    """Convert any date value to YYYY-MM-DD string."""
    if val is None:
        return None
    # Native datetime / date object from openpyxl
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    v = str(val).strip()
    if not v or v in ("NULL", "None", ""):
        return None
    # Already ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}$", v):
        return v
    # "25-Jan-25" / "25-Jan-2025"
    m = re.match(r"^(\d{1,2})-([A-Za-z]{3})-(\d{2,4})$", v)
    if m:
        day, mon_str, yr = int(m.group(1)), m.group(2).upper(), m.group(3)
        MONTHS = {
            m: i
            for i, m in enumerate(
                [
                    "JAN",
                    "FEB",
                    "MAR",
                    "APR",
                    "MAY",
                    "JUN",
                    "JUL",
                    "AUG",
                    "SEP",
                    "OCT",
                    "NOV",
                    "DEC",
                ],
                1,
            )
        }
        mon = MONTHS.get(mon_str)
        if mon:
            year = int(yr) + (2000 if len(yr) == 2 else 0)
            return f"{year}-{mon:02d}-{day:02d}"
    # dateutil fallback
    try:
        from dateutil import parser as dp

        return dp.parse(v, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        pass
    return v  # return as-is if unparseable


def fix_flight_number(val) -> str | None:
    """
    'JL-0062' -> 'JL62'   (remove hyphen, strip leading zeros from numeric part)
    'EK-0236' -> 'EK236'
    'DL-9698' -> 'DL9698'  (no leading zeros to strip)
    Also handles 'JL0062' without hyphen.
    """
    if val is None:
        return None
    v = str(val).strip()
    if not v or v in ("NULL", "None", ""):
        return None
    # Match carrier (2-3 letters/digits) + optional hyphen + numeric part
    m = re.match(r"^([A-Z0-9]{2,3})-?(\d+)$", v.upper())
    if m:
        carrier = m.group(1)
        number = str(int(m.group(2)))  # int() strips leading zeros
        return f"{carrier}{number}"
    return v  # return as-is if pattern doesn't match


def safe_str(val) -> str | None:
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    v = str(val).strip()
    return None if v in ("", "NULL", "None") else v


# ---------------------------------------------------------------------------
# Detect columns dynamically from header row
# ---------------------------------------------------------------------------


def classify_columns(headers: list[str]):
    """
    Returns:
      date_cols      - set of col names that are DepartureDateLocalN
      flight_cols    - set of col names that are FlightNumberN
      other_cols     - everything else (passed through as string)
    """
    date_cols = set()
    flight_cols = set()
    other_cols = []
    for h in headers:
        hl = h.strip().lower()
        if re.match(r"departuredatelocal\d+", hl):
            date_cols.add(h)
        elif re.match(r"flightnumber\d+", hl):
            flight_cols.add(h)
        else:
            other_cols.append(h)
    return date_cols, flight_cols, other_cols


# ---------------------------------------------------------------------------
# Transform one row dict
# ---------------------------------------------------------------------------


def transform_row(r: dict, date_cols: set, flight_cols: set) -> dict:
    out = {}
    for k, v in r.items():
        if k in date_cols:
            out[k] = parse_date(v)
        elif k in flight_cols:
            out[k] = fix_flight_number(v)
        else:
            out[k] = safe_str(v)
    return out


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

print(f"Opening workbook: {INPUT_FILE}")
wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
sheet_names = wb.sheetnames
print(f"Sheets found: {sheet_names}")
wb.close()

con = duckdb.connect(DB_PATH)
con.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")
table_created = False
grand_total = 0

for sheet in sheet_names:
    print(f"\nProcessing sheet: {sheet}")
    wb2 = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb2[sheet]
    rows_iter = ws.iter_rows(values_only=True)

    # Read header
    raw_headers = next(rows_iter)
    headers = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(raw_headers)
    ]

    date_cols, flight_cols, other_cols = classify_columns(headers)
    print(f"  Date cols   : {sorted(date_cols)}")
    print(f"  Flight cols : {sorted(flight_cols)}")
    print(f"  Other cols  : {len(other_cols)}")

    sheet_total = 0

    def flush(batch):
        global table_created
        if not batch:
            return
        transformed = [transform_row(r, date_cols, flight_cols) for r in batch]
        # Build column-oriented dict for DuckDB insert via relation
        cols_all = list(transformed[0].keys())
        data = {c: [row.get(c) for row in transformed] for c in cols_all}

        import polars as pl

        pdf = pl.DataFrame(data, infer_schema_length=len(transformed), strict=False)

        if not table_created:
            con.execute(f"CREATE TABLE {TABLE_NAME} AS SELECT * FROM pdf LIMIT 0")
            globals()["table_created"] = True
        con.execute(f"INSERT INTO {TABLE_NAME} SELECT * FROM pdf")

    batch = []
    for row in rows_iter:
        # Skip completely empty rows
        if all(v is None for v in row):
            continue
        batch.append(dict(zip(headers, row)))
        if len(batch) >= CHUNK_SIZE:
            flush(batch)
            sheet_total += len(batch)
            print(f"  Inserted {sheet_total:,} rows...", end="\r")
            batch = []

    if batch:
        flush(batch)
        sheet_total += len(batch)

    print(f"  Sheet '{sheet}' done — {sheet_total:,} rows.")
    grand_total += sheet_total
    wb2.close()

count = con.execute(f"SELECT COUNT(*) FROM {TABLE_NAME}").fetchone()[0]
cols = con.execute(f"PRAGMA table_info({TABLE_NAME})").fetchdf()
print(f"\n✅ Done! Table '{TABLE_NAME}' in '{DB_PATH}'")
print(f"   Total rows : {count:,}")
print(f"   Columns    : {len(cols)}")
print(f"\n{cols[['name', 'type']].to_string(index=False)}")
con.close()
